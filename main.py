import tkinter.filedialog
from tkinter import *
import re

import os
import json
import openpyxl
from openpyxl.styles.fills import PatternFill
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.cell import get_column_letter

MATH_OPERATORS = ['+', '-', '*', '/', '(', ')', '^', '=']
KEYCODES_LIST = [49, 50, 51, 52, 53, 54, 55, 56, 57, 48, 81, 87, 69, 82, 84, 89, 85, 73, 79, 80, 65,
                 83, 68, 70, 71, 72, 74, 75, 76, 90, 88, 67, 86, 66, 78, 77]
BAD_FILE_NAME_CHARS = ['\\', '/', '*', ':', '"', '?', '|', '<', '>']
global root, counters_list, subtraction_flag, CONFIG, inner_counters_names_dict, inner_counters_dict, score_box, \
    counters_per_column
HEADER_FONT = ("Calibri Light", 14)
TEXT_FONT = ("Calibri Light", 10)
BACKGROUND_COLOR = 'lightblue1'


def get_formula(formula_string: str, table_index: str, player_row: int = 2):
    formula = formula_string.split("=")
    name = formula[0]
    formula = "=" + formula[1]
    for counter in CONFIG["player counters names list"][::-1] + CONFIG["general counters names list"][::-1]:
        counter_indexes_list = [m.start() for m in re.finditer(counter, formula)]
        for substring_index in counter_indexes_list[::-1]:
            if substring_index == 1 or substring_index + len(counter) == len(formula) or (formula[substring_index - 1] in MATH_OPERATORS and formula[substring_index + len(counter)] in MATH_OPERATORS):
                formula = formula[:substring_index] + f'HLOOKUP("{counter}",{table_index},{player_row},FALSE)' \
                          + formula[substring_index + len(counter):]
    return [name, formula]


class Counter:
    def __init__(self, window, row: int, column: int, name="empty", count=0):
        self.name = name
        self.count = count
        self.row = row
        self.column = column
        self.window = window


class EntryCounter(Counter):
    def __init__(self, window, row: int, column: int, name="empty", count=0):
        super().__init__(window, row, column, name, count)
        self.entry = Entry(window, width=8, borderwidth=2)


class KeysCounter(Counter):
    def __init__(self, window, key: int, row: int, column: int, name="empty", count=0):
        super().__init__(window, row, column, name, count)
        self.key = key
        self.label = None
        self.display()

    def add_or_subtract_one(self, add_or_subtract_flag):
        if add_or_subtract_flag:
            self.count -= 1
        else:
            self.count += 1
        self.display()

    def display(self):
        Label(self.window, text=f"{chr(self.key).lower()}", font=TEXT_FONT, pady=1, bg="lightblue1") \
            .grid(row=self.row, column=self.column)
        Label(self.window, text=f"{self.name}", font=TEXT_FONT, pady=1, bg="lightblue1") \
            .grid(row=self.row, column=self.column + 1)
        self.label = Label(self.window, text=f"{self.count}", font=TEXT_FONT, pady=1, bg="lightblue1")
        self.label.grid(row=self.row, column=self.column + 2)


def init_root_screen(filename: str):
    global counters_list, subtraction_flag, inner_counters_names_dict, inner_counters_dict, counters_per_column
    root.configure(bg="lightblue1")
    counters_per_column = 22
    '''
    background_image = ImageTk.PhotoImage(file="background_image.jpg")
    background_label = Label(root,image=background_image)
    background_label.image=background_image
    background_label.grid(row=0,column=0)
    '''  # adds background image, currently not working
    counters_list = []
    subtraction_flag = False
    for i in range(len(CONFIG["player counters names list"])):
        counters_list.append(
            KeysCounter(root, KEYCODES_LIST[i], int(i - int(i / counters_per_column) * counters_per_column) + 3,
                        int(i / counters_per_column) * 3,
                        CONFIG["player counters names list"][i]))
    for i in range(int((len(counters_list) - 1) / counters_per_column) + 1):
        Label(root, text="Key", font=TEXT_FONT, bg=BACKGROUND_COLOR, fg="black", bd=3) \
            .grid(row=2, column=i * 3)
        Label(root, text="Content", font=TEXT_FONT, bg=BACKGROUND_COLOR, fg="black", bd=3) \
            .grid(row=2, column=i * 3 + 1)
        Label(root, text="Count", font=TEXT_FONT, bg=BACKGROUND_COLOR, fg="black", bd=3) \
            .grid(row=2, column=i * 3 + 2)
    Label(root, text=filename, font=HEADER_FONT, bg=BACKGROUND_COLOR, padx=30, bd=3) \
        .grid(row=0, column=0, columnspan=(int(len(counters_list) / counters_per_column) + 1) * 3)
    Label(root, text="Press '-' To Subtract", font=TEXT_FONT, bg=BACKGROUND_COLOR, padx=30, bd=3) \
        .grid(row=1, column=0, columnspan=(int(len(counters_list) / counters_per_column) + 1) * 3)
    export_button = Button(root, text="Export", font=TEXT_FONT, bg='snow', padx=30, bd=3,
                           command=export_to_excel)

    export_button.grid(row=counters_per_column + 5, column=int((len(counters_list) - 1) / counters_per_column) * 3,
                       sticky=E, columnspan=3)

    root.resizable(False, False)
    root.eval('tk::PlaceWindow . center')
    inner_counters_dict = {}
    if CONFIG["MAIN CONFIG"]:
        inner_counters_names_dict = {}
        for counter in CONFIG["player counters names list"]:
            with open(f"{os.getcwd()}/configurations/{counter}.json") as file:
                counter_config = json.load(file)
                inner_counters_names_dict[counter] = counter_config["player counters names list"]
    else:
        import_button = Button(root, text="Import", font=TEXT_FONT, bg='snow', padx=30, bd=3,
                               command=import_from_excel)
        import_button.grid(row=counters_per_column + 5, column=0, sticky=W, columnspan=3)


def key_pressed_root_window(event):
    global subtraction_flag
    if event.char == '-':
        subtraction_flag = not subtraction_flag
    for i in range(len(counters_list)):
        if event.keycode == KEYCODES_LIST[i]:
            counters_list[i].add_or_subtract_one(subtraction_flag)
            if CONFIG["MAIN CONFIG"] and not subtraction_flag:
                init_attack_counter_window(counters_list[i])


def key_pressed_inside_attack_window(event, counter_name):
    global subtraction_flag
    if event.char == '-':
        subtraction_flag = not subtraction_flag
    for i in inner_counters_dict[counter_name]:
        if event.keycode == i.key:
            i.add_or_subtract_one(subtraction_flag)


def withdraw_window_and_focus_root(window):
    window.withdraw()
    root.focus()


def init_attack_counter_window(counter):
    global inner_counters_dict
    if counter.name in inner_counters_dict.keys():
        inner_counters_dict[counter.name][0].window.deiconify()
    else:
        counter_window = Toplevel(root)
        inner_counters_dict[counter.name] = []
        for c_index in range(len(inner_counters_names_dict[counter.name])):
            inner_counters_dict[counter.name].append(KeysCounter(counter_window, KEYCODES_LIST[c_index], c_index + 3, 0,
                                                                 inner_counters_names_dict[counter.name][c_index]))
        for i in range(int((len(inner_counters_dict[counter.name]) - 1) / counters_per_column) + 1):
            Label(counter_window, text="Key", font=TEXT_FONT, bg=BACKGROUND_COLOR, fg="black", bd=3) \
                .grid(row=2, column=i * 3)
            Label(counter_window, text="Content", font=TEXT_FONT, bg=BACKGROUND_COLOR, fg="black", bd=3) \
                .grid(row=2, column=i * 3 + 1)
            Label(counter_window, text="Count", font=TEXT_FONT, bg=BACKGROUND_COLOR, fg="black", bd=3) \
                .grid(row=2, column=i * 3 + 2)
        Label(counter_window, text=counter.name, font=HEADER_FONT, bg=BACKGROUND_COLOR, padx=30, bd=3) \
            .grid(row=0, column=0, columnspan=(int(len(counters_list) / counters_per_column) + 1) * 3)
        Label(counter_window, text="Press '-' To Subtract", font=TEXT_FONT, bg=BACKGROUND_COLOR, padx=30, bd=3) \
            .grid(row=1, column=0, columnspan=(int(len(counters_list) / counters_per_column) + 1) * 3)
        counter_window.title = counter.name
        counter_window.bind('<Key>', lambda event: key_pressed_inside_attack_window(event, counter.name))
        counter_window.bind('<Return>', lambda event: withdraw_window_and_focus_root(counter_window))
        counter_window.configure(bg=BACKGROUND_COLOR)
        counter_window.attributes('-disabled', True)
        counter_window.geometry("+100+100")
        counter_window.focus()


def isfloat(string):
    try:
        float(string)
        return True
    except ValueError:
        return False


def save_and_close(file_name, player_name, window_to_destroy):
    flag = True
    for counter in score_box:
        if isfloat(counter.entry.get()):
            if "." in counter.entry.get():
                counter.count = float(counter.entry.get())
            else:
                counter.count = int(counter.entry.get())
        else:
            flag = False
            break
    if file_name != '' and not any([c in BAD_FILE_NAME_CHARS for c in file_name]) and flag:
        r = Rule(type="expression", dxf=DifferentialStyle(fill=PatternFill(bgColor="FFC7CE")), stopIfTrue=True)
        r.formula = ['A$2>0']
        file_full_location = rf"{CONFIG['SAVE LOCATION']}\{file_name}.xlsx"
        export_results = openpyxl.Workbook()
        main_sheet = export_results.active
        main_sheet.title = "Main"
        columns = ["player name"] + [i.name for i in score_box] + [i.name for i in counters_list]
        data = [player_name] + [i.count for i in score_box] + [i.count for i in counters_list]
        main_sheet.append(columns)
        main_sheet.append(data)
        main_table_size = main_sheet.dimensions
        for formula_string_index in range(len(CONFIG["ADVANCED STATS"])):
            temp_formula = get_formula(CONFIG["ADVANCED STATS"][formula_string_index], main_table_size)
            main_sheet[f"{get_column_letter(formula_string_index + 1)}4"] = temp_formula[0]
            main_sheet[f"{get_column_letter(formula_string_index + 1)}5"] = temp_formula[1]
        main_sheet.conditional_formatting.add("A2:GZ2", r)
        for attack, counters in inner_counters_dict.items():
            columns = [i.name for i in counters]
            data = [i.count for i in counters]
            new_sheet = export_results.create_sheet(attack)
            new_sheet.append(columns)
            new_sheet.append(data)
            table_size = new_sheet.dimensions
            with open(f"{os.getcwd()}/configurations/{attack}.json") as file:
                calculations = json.load(file)["ADVANCED STATS"]
                for formula_string_index in range(len(calculations)):
                    temp_formula = get_formula(calculations[formula_string_index], table_size)
                    new_sheet[f"{get_column_letter(formula_string_index + 1)}4"] = temp_formula[0]
                    new_sheet[f"{get_column_letter(formula_string_index + 1)}5"] = temp_formula[1]
            new_sheet.conditional_formatting.add("A2:AZ2", r)
        export_results.save(file_full_location)
        window_to_destroy.destroy()


def export_to_excel():
    global score_box
    export_window = Toplevel(root)
    export_window.resizable(False, False)
    export_window.title = "name the file"
    Label(export_window, text="file name:", font=TEXT_FONT).grid(row=0, column=0, columnspan=2)
    file_name_entry = Entry(export_window, width=35, borderwidth=5)
    file_name_entry.grid(row=0, column=2, columnspan=3)
    file_name_entry.focus()
    Label(export_window, text="player name:", font=TEXT_FONT).grid(row=1, column=0, columnspan=2)
    player_name_entry = Entry(export_window, width=20, borderwidth=3)
    player_name_entry.grid(row=1, column=2, columnspan=2)
    Label(export_window, text="", font=TEXT_FONT).grid(row=2, column=0)
    score_box = []
    for i in range(len(CONFIG["general counters names list"])):
        score_box.append(
            EntryCounter(export_window, i % 16 + 3, 1 + int(i / 16) * 2, CONFIG["general counters names list"][i]))
        score_box[i].entry.grid(row=score_box[i].row, column=score_box[i].column)
        score_box[i].entry.insert(-1, "0")
        Label(export_window, text=f'{CONFIG["general counters names list"][i]}:', font=("Calibri Light", 12)) \
            .grid(row=i % 16 + 3, column=+int(i / 16) * 2)
    Button(export_window, text="done", font=TEXT_FONT,
           command=lambda: save_and_close(file_name_entry.get(), player_name_entry.get(), export_window)) \
        .grid(row=1200, column=2, columnspan=2)
    export_window.bind('<Return>',
                       lambda event: save_and_close(file_name_entry.get(), player_name_entry.get(), export_window))


def import_from_excel(override_statistics=False, filename=None, old_window=None):
    if not override_statistics:
        filename = tkinter.filedialog.askopenfilename(filetypes=[("Excel file", "*.xlsx")],
                                                      initialdir=CONFIG["SAVE LOCATION"])
    new_window = Toplevel(root)
    new_window.title = "calculation finished"
    new_window.resizable(False, False)
    new_window.geometry(f"+{int(root.winfo_screenwidth() / 2)}+{int(root.winfo_screenheight() / 2)}")
    if filename != '':
        book = openpyxl.load_workbook(filename)
        if "statistics" not in book or override_statistics:
            import_results = openpyxl.load_workbook(filename)
            data_sheet = import_results.active
            if old_window:
                old_window.destroy()
                import_results.remove(import_results["statistics"])
            statistics_sheet = import_results.create_sheet("statistics")
            for row_index in range(1, data_sheet.max_row + 1):
                statistics_sheet[f"A{row_index}"] = f'=HLOOKUP("player name",{data_sheet.title}!' \
                                                    f'{data_sheet.dimensions},{row_index},FALSE)'
            for formula_string_index in range(len(CONFIG["ADVANCED STATS"])):
                temp_formula = ["", ""]
                for row_index in range(2, data_sheet.max_row + 1):
                    temp_formula = get_formula(CONFIG["ADVANCED STATS"][formula_string_index],
                                               f"{data_sheet.title}!{data_sheet.dimensions}", row_index)
                    statistics_sheet[f"{get_column_letter(formula_string_index + 2)}{row_index}"] = temp_formula[1]
                statistics_sheet[f"{get_column_letter(formula_string_index + 2)}1"] = temp_formula[0]

            import_results.save(filename)
            Label(new_window, text="Finished! everything worked", padx=30, pady=15).pack()
            Button(new_window, text="close", command=new_window.destroy).pack()
        else:
            Label(new_window, text="want to override statistics?", padx=30, pady=15).pack()
            Button(new_window, padx=20, bd=3, text="yes",
                   command=lambda: import_from_excel(True, filename, new_window)).pack(padx=20, side=LEFT)
            Button(new_window, padx=20, bd=3, text="no", command=new_window.destroy).pack(padx=20, side=RIGHT)
    else:
        Label(new_window, text="didn't choose a file", padx=30, pady=15).pack()
        Button(new_window, text="close", command=new_window.destroy).pack()


def main():
    global root, CONFIG
    filename = tkinter.filedialog.askopenfilename(filetypes=[("Json File", "*.json")],
                                                  initialdir=os.getcwd() + "/configurations")
    root = Tk()
    with open(filename) as file:
        CONFIG = json.load(file)
        init_root_screen(filename.split("/")[-1].split(".")[0])
        root.bind('<Key>', key_pressed_root_window)
        root.mainloop()


if __name__ == "__main__":
    main()
